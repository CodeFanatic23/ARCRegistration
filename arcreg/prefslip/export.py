import csv
import xlwt
from django.utils.encoding import smart_str
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.http import HttpResponseRedirect

class ExportAddCourses:
    def export_csv(self,modeladmin, request, queryset):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename=ADD_ELEC_PREF_CASES.csv'
        writer = csv.writer(response, csv.excel)
        response.write(u'\ufeff'.encode('utf8')) # BOM (optional...Excel needs it to open UTF-8 file properly)
        writer.writerow([
            smart_str(u"ID"),
            smart_str(u"Name"),
            smart_str(u"ID_no"),
            smart_str(u"Course ID"),
            smart_str(u"Class No."),
            smart_str(u"Course Title"),
            smart_str(u"Priority"),
            smart_str(u"Priority Number"),
        ])
        for obj in queryset:
            writer.writerow([
                smart_str(obj.pk),
                smart_str(obj.name),
                smart_str(obj.ID_no),
                smart_str(obj.course_id),
                smart_str(obj.class_nbr),
                smart_str(obj.course_title),
                smart_str(obj.priority),
                smart_str(obj.pr_no),
            ])
        return response


    def export_xls(self,modeladmin, request, queryset):
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename=ADD_ELEC_PREF_CASES.xls'
        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet("ADD_ELEC_PREF_CASES")

        row_num = 0

        columns = [
            (u"ID", 2000),
            (u"Name", 6000),
            (u"ID No.", 8000),
            (u"Course ID", 8000),
            (u"Class No.", 8000),
            (u"Course Title", 8000),
            (u"Priority", 8000),
            (u"Priority Number", 8000),
        ]

        font_style = xlwt.XFStyle()
        font_style.font.bold = True

        for col_num in range(len(columns)):
            ws.write(row_num, col_num, columns[col_num][0], font_style)
            # set column width
            ws.col(col_num).width = columns[col_num][1]

        font_style = xlwt.XFStyle()
        font_style.alignment.wrap = 1

        for obj in queryset:
            row_num += 1
            row = [
                obj.pk,
                obj.name,
                obj.ID_no,
                obj.course_id,
                obj.class_nbr,
                obj.course_title,
                obj.priority,
                obj.pr_no,
            ]
            for col_num in range(len(row)):
                ws.write(row_num, col_num, row[col_num], font_style)
                
        wb.save(response)
        return response

    def export_xlsx(self,modeladmin, request, queryset):
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=ADD_ELEC_PREF_CASES.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.get_active_sheet()
        ws.title = "ADD_ELEC_PREF_CASES"

        row_num = 0

        columns = [
            (u"ID", 15),
            (u"Name", 70),
            (u"ID No.", 70),
            (u"Course ID", 70),
            (u"Class No.", 70),
            (u"Course Title", 70),
            (u"Priority", 70),
            (u"Priority Number", 70),
        ]

        for col_num in range(len(columns)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = columns[col_num][0]
            c.font = c.font.copy(bold = True)
            # set column width
            ws.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num][1]

        for obj in queryset:
            row_num += 1
            row = [
                obj.pk,
                obj.name,
                obj.ID_no,
                obj.course_id,
                obj.class_nbr,
                obj.course_title,
                obj.priority,
                obj.pr_no,
            ]
            for col_num in range(len(row)):
                c = ws.cell(row=row_num + 1, column=col_num + 1)
                c.value = row[col_num]
                c.alignment = c.alignment.copy(wrap_text = True)

        wb.save(response)
        return response