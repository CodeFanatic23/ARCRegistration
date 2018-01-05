from django.contrib import admin
from django.http import HttpResponse
from django.http import HttpResponseRedirect
from django.core.urlresolvers import reverse
from .models import *
from .forms import *
from upload_sheet.models import Time_Table_Semester_Wise
import csv
import xlwt
from django.utils.encoding import smart_str
import openpyxl
from openpyxl.utils import get_column_letter
from .export import *

# Register your models here.
class RemoveCasesAdmin(admin.ModelAdmin):
	list_display=['name','ID_no','course_no','course_id','class_nbr','course_title','Section']
	search_fields = ['name','ID_no','course_title','class_nbr']
	#actions = ['export_xls','export_xlsx','export_csv']

	def Section(self,obj):
		return 'L' + obj.lecture_no + ' T' + obj.tutorial_no + ' P' + obj.practical_no + ' ' + obj.graded_comp + ' ' + obj.project_section
	# For refactoring
	rem = ExportRemove()
	funcs = {'export_xls':rem.export_csv,'export_csv':rem.export_xlsx,'export_xlsx':rem.export_xls}
	actions = [funcs['export_xls'],funcs['export_xlsx'],funcs['export_csv']]
	
	
	model=Remove

admin.site.register(Remove,RemoveCasesAdmin)

class AddCasesAdmin(admin.ModelAdmin):
	list_display=['name','ID_no','course_no','course_id','class_nbr','course_title','Section']
	search_fields = ['name','ID_no','course_title','class_nbr']
	
	def Section(self,obj):
		return 'L' + obj.lecture_no + ' T' + obj.tutorial_no + ' P' + obj.practical_no + ' ' + obj.graded_comp + ' ' + obj.project_section
	add = ExportAdd()
	funcs = {'export_xls':add.export_csv,'export_csv':add.export_xlsx,'export_xlsx':add.export_xls}
	actions = [funcs['export_xls'],funcs['export_xlsx'],funcs['export_csv']]
	
	
	model=Add

admin.site.register(Add,AddCasesAdmin)


class UserDataAdmin(admin.ModelAdmin):
	list_display=['name','ID_no','phone_no','submit','message']
	list_editable=['message']
	search_fields = ["name","ID_no"]
	
	model=Registered_User
	def save_model(self, request, obj, form, change=False):
		print("hello before saving")
		obj.message_status = False
		obj.save()
		print("Saved!")

	def submit(self,obj):
		return obj.submit_status
		

admin.site.register(Registered_User,UserDataAdmin)

class InstructionAdmin(admin.ModelAdmin):
	list_display=['instruction']
	list_display_links=['instruction']
	model = Instruction

admin.site.register(Instruction,InstructionAdmin)

class Generate_UserAdmin(admin.ModelAdmin):
	list_display = ['username','password']

	form = Generate_UserForm
	model = Generate_User

	def username(self,obj):
		return obj.no_of_users_to_generate

	def password(self,obj):
		return obj.username_pattern
	
	def response_add(self, request, obj, post_url_continue=None):
		return HttpResponseRedirect(reverse("admin:blueslip_generated_user_changelist"))

	# def response_add(self, request, obj, post_url_continue=None):
 #    """This makes the response go to the newly created model's change page
 #    without using reverse"""
 #    return HttpResponseRedirect("../%s" % obj.id])

class Generated_UserAdmin(admin.ModelAdmin):
	list_display_links=None
	list_display=['username','password']
	actions=['export_xls','export_xlsx','export_csv']

	class Media:
		js = ('/static/actions.min.js','http://forms.viewflow.io/static/material/admin/js/admin_init.js')
	def username(self,obj):
		return obj.usrname

	def password(self,obj):
		return obj.pwd

	def has_add_permission(self,request):
		return False

	def export_xls(modeladmin, request, queryset):
		response = HttpResponse(content_type='application/ms-excel')
		response['Content-Disposition'] = 'attachment; filename=BULK_USERS.xls'
		wb = xlwt.Workbook(encoding='utf-8')
		ws = wb.add_sheet("BULK_USERS")

		row_num = 0

		columns = [
		    (u"ID", 2000),
		    (u"Username", 6000),
		    (u"Password", 8000),
		]

		font_style = xlwt.XFStyle()
		font_style.font.bold = True

		for col_num in range(len(columns)):
		    ws.write(row_num, col_num, columns[col_num][0], font_style)
		    # set column width
		    ws.col(col_num).width = columns[col_num][1]

		font_style = xlwt.XFStyle()
		font_style.alignment.wrap = 1

		for obj in queryset:
		    row_num += 1
		    row = [
		        obj.pk,
		        obj.usrname,
		        obj.pwd,
		    ]
		    for col_num in range(len(row)):
		        ws.write(row_num, col_num, row[col_num], font_style)
		        
		wb.save(response)
		return response

	export_xls.short_description = u"Export XLS"
	def export_xlsx(modeladmin, request, queryset):
		response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
		response['Content-Disposition'] = 'attachment; filename=BULK_USERS.xlsx'
		wb = openpyxl.Workbook()
		ws = wb.get_active_sheet()
		ws.title = "BULK_USERS"

		row_num = 0

		columns = [
		    (u"ID", 15),
		    (u"Username", 70),
		    (u"Password", 70),
		]

		for col_num in range(len(columns)):
		    c = ws.cell(row=row_num + 1, column=col_num + 1)
		    c.value = columns[col_num][0]
		    c.font = c.font.copy(bold = True)
		    # set column width
		    ws.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num][1]

		for obj in queryset:
		    row_num += 1
		    row = [
		        obj.pk,
		        obj.usrname,
		        obj.pwd,
		      
		    ]
		    for col_num in range(len(row)):
		        c = ws.cell(row=row_num + 1, column=col_num + 1)
		        c.value = row[col_num]
		        c.alignment = c.alignment.copy(wrap_text = True)

		wb.save(response)
		return response

	export_xlsx.short_description = u"Export XLSX"

	def export_csv(modeladmin, request, queryset):
		response = HttpResponse(content_type='text/csv')
		response['Content-Disposition'] = 'attachment; filename=BULK_USERS.csv'
		writer = csv.writer(response, csv.excel)
		response.write(u'\ufeff'.encode('utf8')) # BOM (optional...Excel needs it to open UTF-8 file properly)
		writer.writerow([
		    smart_str(u"ID"),
		    smart_str(u"Username"),
		    smart_str(u"Password"),
		])
		for obj in queryset:
		    writer.writerow([
		        smart_str(obj.pk),
		        smart_str(obj.usrname),
		        smart_str(obj.pwd),
		    ])
		return response
	export_csv.short_description = u"Export CSV"

class ControlAdmin(admin.ModelAdmin):
	list_display = ['__str__']

	model = Control

	def has_add_permission(self,request):
		if self.model.objects.count() >= 1:
			return False
		else:
			return True
class CourseSelectorAdmin(admin.ModelAdmin):
	list_display = ['discipline']
	model = Discipline

	
admin.site.register(Discipline,CourseSelectorAdmin)
admin.site.register(Control,ControlAdmin)

admin.site.register(Generated_User,Generated_UserAdmin)
admin.site.register(Generate_User,Generate_UserAdmin)
